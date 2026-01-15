from fastapi import FastAPI, Depends, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from jose import jwt, JWTError
import bcrypt
from datetime import datetime, timedelta, date, time
from typing import Optional
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

from database import SessionLocal, engine, Base
from models import User, Event, EventRegistration, EventRequest
from schemas import (
    UserCreate, UserResponse, UserUpdate,
    EventCreate, EventResponse, EventUpdate,
    EventRegistrationCreate, EventRegistrationResponse,
    EventRequestCreate, EventRequestResponse,
    Token, DescriptionGenerateRequest, DescriptionGenerateResponse
)

# Create tables
Base.metadata.create_all(bind=engine)

app = FastAPI()

# CORS - Allow all origins for network access
# In production, you should restrict this to specific domains
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins for network access
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Security
SECRET_KEY = os.getenv("SECRET_KEY", "your-secret-key-change-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/login")

# Database dependency
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Password hashing using bcrypt directly
def verify_password(plain_password: str, hashed_password: str) -> bool:
    """Verify a password against a hash"""
    try:
        # Если хеш начинается с $2b$ или $2a$, это bcrypt
        if hashed_password.startswith('$2'):
            return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))
        # Если это старый pbkdf2 хеш, возвращаем False
        return False
    except Exception as e:
        print(f"Password verification error: {e}")
        return False

def get_password_hash(password: str) -> str:
    """Hash a password using bcrypt"""
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password.encode('utf-8'), salt)
    return hashed.decode('utf-8')

# JWT
def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

# Get current user
def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Тіркелгілерді тексеру мүмкін болмады",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        # Manual token extraction
        if not token:
            raise credentials_exception
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    user = db.query(User).filter(User.id == int(user_id)).first()
    if user is None:
        raise credentials_exception
    return user

# Auth endpoints
@app.post("/api/register", response_model=UserResponse)
def register(user: UserCreate, db: Session = Depends(get_db)):
    # Check if email already exists
    db_user = db.query(User).filter(User.email == user.email).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Email қазірдің өзінде тіркелген")
    
    # Check if admin registration
    if user.role == "admin":
        if not user.secret_code or user.secret_code != "111111":
            raise HTTPException(status_code=403, detail="Администратор тіркелуі үшін құпия код дұрыс емес")
    
    hashed_password = get_password_hash(user.password)
    db_user = User(
        email=user.email,
        hashed_password=hashed_password,
        full_name=user.full_name,
        group=user.group,
        role=user.role
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user

@app.post("/api/login", response_model=Token)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    try:
        user = db.query(User).filter(User.email == form_data.username).first()
        if not user:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Дұрыс емес email немесе құпия сөз",
                headers={"WWW-Authenticate": "Bearer"},
            )
        
        # Проверяем пароль с обработкой ошибок
        try:
            password_valid = verify_password(form_data.password, user.hashed_password)
        except Exception as e:
            print(f"Password verification error: {str(e)}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Құпия сөзді тексеру сәтсіз аяқталды: {str(e)}"
            )
        
        if not password_valid:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Дұрыс емес email немесе құпия сөз",
                headers={"WWW-Authenticate": "Bearer"},
            )
        
        access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        access_token = create_access_token(
            data={"sub": str(user.id)}, expires_delta=access_token_expires
        )
        return {"access_token": access_token, "token_type": "bearer"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Login error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Internal server error: {str(e)}"
        )

@app.get("/api/me", response_model=UserResponse)
def read_users_me(current_user: User = Depends(get_current_user)):
    return current_user

@app.put("/api/users/me", response_model=UserResponse)
def update_user_profile(
    user_update: UserUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update current user's profile"""
    try:
        if user_update.full_name is not None:
            current_user.full_name = user_update.full_name
        if user_update.group is not None:
            current_user.group = user_update.group
        if user_update.photo_url is not None:
            # Validate base64 string length (max 10MB image as base64 ~13.3MB)
            if len(user_update.photo_url) > 14000000:
                raise HTTPException(
                    status_code=400,
                    detail="Фото слишком большое. Максимальный размер: 10MB"
                )
            current_user.photo_url = user_update.photo_url
        
        db.commit()
        db.refresh(current_user)
        return current_user
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка при обновлении профиля: {str(e)}"
        )

# Event endpoints
@app.get("/api/events", response_model=list[EventResponse])
def read_events(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    events = db.query(Event).offset(skip).limit(limit).all()
    return events

@app.get("/api/events/{event_id}", response_model=EventResponse)
def read_event(event_id: int, db: Session = Depends(get_db)):
    event = db.query(Event).filter(Event.id == event_id).first()
    if event is None:
        raise HTTPException(status_code=404, detail="Іс-шара табылмады")
    return event

@app.post("/api/events", response_model=EventResponse)
def create_event(event: EventCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Тек администраторлар іс-шара құра алады")
    
    # Validate max_participants limit
    if event.max_participants > 200:
        raise HTTPException(
            status_code=400,
            detail="Максимум 200 қатысушыға рұқсат етілген"
        )
    
    # Validate image_url length if provided
    if event.image_url and len(event.image_url) > 14000000:
        raise HTTPException(
            status_code=400,
            detail="Фото слишком большое. Максимальный размер: 10MB"
        )
    
    try:
        # Prepare event data, excluding image_url if it's empty string
        event_data = event.dict()
        if event_data.get('image_url') == '':
            event_data['image_url'] = None
        
        db_event = Event(
            **event_data,
            created_by=current_user.id
        )
        db.add(db_event)
        db.commit()
        db.refresh(db_event)
        return db_event
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error creating event: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Іс-шараны құру кезінде қате пайда болды: {str(e)}"
        )

@app.put("/api/events/{event_id}", response_model=EventResponse)
def update_event(
    event_id: int,
    event: EventUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Тек администраторлар іс-шараны жаңарта алады")
    
    db_event = db.query(Event).filter(Event.id == event_id).first()
    if not db_event:
        raise HTTPException(status_code=404, detail="Іс-шара табылмады")
    
    update_data = event.dict(exclude_unset=True)
    
    # Validate max_participants limit
    if 'max_participants' in update_data and update_data['max_participants'] is not None:
        if update_data['max_participants'] > 200:
            raise HTTPException(
                status_code=400,
                detail="Максимум 200 қатысушыға рұқсат етілген"
            )
    
    # Validate image_url length if provided
    if 'image_url' in update_data and update_data['image_url'] is not None:
        if update_data['image_url'] == '':
            update_data['image_url'] = None
        elif len(update_data['image_url']) > 14000000:
            raise HTTPException(
                status_code=400,
                detail="Фото слишком большое. Максимальный размер: 10MB"
            )
    
    try:
        for field, value in update_data.items():
            if value is None:
                continue # Skip None values
                
            if field == 'date':
                # Convert string date to date object
                from datetime import date as dt_date
                try:
                    date_obj = dt_date.fromisoformat(value)
                    setattr(db_event, field, date_obj)
                except ValueError:
                    raise HTTPException(status_code=400, detail=f"Неверный формат даты для '{field}': {value}. Ожидается YYYY-MM-DD.")
            elif field == 'start_time':
                # Convert string time to time object
                from datetime import datetime as dt
                try:
                    time_obj = dt.strptime(value, '%H:%M').time()
                    setattr(db_event, field, time_obj)
                except ValueError:
                    try:
                        time_obj = dt.strptime(value, '%H:%M:%S').time()
                        setattr(db_event, field, time_obj)
                    except ValueError:
                        raise HTTPException(status_code=400, detail=f"Неверный формат времени для '{field}': {value}. Ожидается HH:MM или HH:MM:SS.")
            else:
                setattr(db_event, field, value)
        
        db.commit()
        db.refresh(db_event)
        return db_event
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error updating event: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Іс-шараны жаңарту кезінде қате пайда болды: {str(e)}"
        )

@app.delete("/api/events/{event_id}")
def delete_event(event_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Тек администраторлар іс-шараны жоя алады")
    
    db_event = db.query(Event).filter(Event.id == event_id).first()
    if not db_event:
        raise HTTPException(status_code=404, detail="Іс-шара табылмады")
    
    db.delete(db_event)
    db.commit()
    return {"message": "Event deleted"}

# Event registration endpoints
@app.post("/api/events/{event_id}/register", response_model=EventRegistrationResponse)
def register_for_event(
    event_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Prevent admin from registering
    if current_user.role == "admin":
        raise HTTPException(status_code=403, detail="Администраторлар іс-шараға тіркеле алмайды")
    
    # Check if event exists
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Іс-шара табылмады")
    
    # Check if event time has passed
    event_datetime = datetime.combine(event.date, event.start_time)
    if datetime.now() >= event_datetime:
        raise HTTPException(status_code=400, detail="Өткен іс-шараларға тіркелу мүмкін емес")
    
    # Check if already registered
    existing_registration = db.query(EventRegistration).filter(
        EventRegistration.event_id == event_id,
        EventRegistration.user_id == current_user.id
    ).first()
    if existing_registration:
        raise HTTPException(status_code=400, detail="Сіз бұл іс-шараға қазірдің өзінде тіркелгенсіз")
    
    # Check if event is full
    registrations_count = db.query(EventRegistration).filter(EventRegistration.event_id == event_id).count()
    if registrations_count >= event.max_participants:
        raise HTTPException(status_code=400, detail="Іс-шара толып қалды")
    
    registration = EventRegistration(
        event_id=event_id,
        user_id=current_user.id
    )
    db.add(registration)
    
    # Award points (10 points per event) - only for students
    if current_user.role == "student":
        if current_user.points is None:
            current_user.points = 0
        current_user.points += 10
    
    db.commit()
    db.refresh(registration)
    return registration

@app.get("/api/events/{event_id}/is-registered")
def check_registration(event_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    registration = db.query(EventRegistration).filter(
        EventRegistration.event_id == event_id,
        EventRegistration.user_id == current_user.id
    ).first()
    return {"is_registered": registration is not None}

@app.get("/api/events/{event_id}/stats")
def get_event_stats(event_id: int, db: Session = Depends(get_db)):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Іс-шара табылмады")
    
    registrations_count = db.query(EventRegistration).filter(EventRegistration.event_id == event_id).count()
    available_spots = event.max_participants - registrations_count
    is_full = available_spots <= 0
    
    return {
        "current_registrations": registrations_count,
        "max_participants": event.max_participants,
        "available_spots": available_spots,
        "is_full": is_full
    }

@app.get("/api/my-events", response_model=list[EventResponse])
def get_my_events(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    registrations = db.query(EventRegistration).filter(EventRegistration.user_id == current_user.id).all()
    event_ids = [reg.event_id for reg in registrations]
    events = db.query(Event).filter(Event.id.in_(event_ids)).all()
    return events

# Calendar endpoint
@app.get("/api/calendar")
def get_calendar(year: int, month: int, db: Session = Depends(get_db)):
    """
    Get calendar data for a specific month and year.
    Returns events grouped by day.
    """
    from calendar import month_name, monthrange
    from datetime import date as dt_date
    
    # Get all events for the specified month
    start_date = dt_date(year, month, 1)
    if month == 12:
        end_date = dt_date(year + 1, 1, 1)
    else:
        end_date = dt_date(year, month + 1, 1)
    
    events = db.query(Event).filter(
        Event.date >= start_date,
        Event.date < end_date
    ).all()
    
    # Group events by date
    days_dict = {}
    for event in events:
        event_date_str = event.date.isoformat()
        if event_date_str not in days_dict:
            days_dict[event_date_str] = []
        days_dict[event_date_str].append({
            "id": event.id,
            "title": event.title,
            "description": event.description,
            "date": event.date.isoformat(),
            "start_time": event.start_time.strftime("%H:%M:%S"),
            "location": event.location,
            "max_participants": event.max_participants,
            "created_by": event.created_by
        })
    
    # Create days list
    days_list = []
    _, last_day = monthrange(year, month)
    
    for day in range(1, last_day + 1):
        event_date = dt_date(year, month, day)
        event_date_str = event_date.isoformat()
        days_list.append({
            "date": event_date_str,
            "events": days_dict.get(event_date_str, [])
        })
    
    # Get month name in Kazakh
    months_kz = [
        'Қаңтар', 'Ақпан', 'Наурыз', 'Сәуір', 'Мамыр', 'Маусым',
        'Шілде', 'Тамыз', 'Қыркүйек', 'Қазан', 'Қараша', 'Желтоқсан'
    ]
    
    return {
        "year": year,
        "month": month,
        "month_name": months_kz[month - 1],
        "days": days_list
    }

# Event request endpoints
@app.post("/api/event-requests", response_model=EventRequestResponse)
def create_event_request(
    request: EventRequestCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Validate max_participants limit
    if request.max_participants > 200:
        raise HTTPException(
            status_code=400,
            detail="Максимум 200 қатысушыға рұқсат етілген"
        )
    
    event_request = EventRequest(
        **request.dict(),
        user_id=current_user.id,
        status="pending"
    )
    db.add(event_request)
    db.commit()
    db.refresh(event_request)
    return event_request

@app.get("/api/event-requests", response_model=list[EventRequestResponse])
def get_event_requests(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Тек администраторлар барлық өтінімдерді көре алады")
    
    # Sort by created_at descending (newest first)
    requests = db.query(EventRequest).order_by(EventRequest.created_at.desc()).all()
    return requests

@app.get("/api/my-event-requests", response_model=list[EventRequestResponse])
def get_my_event_requests(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    requests = db.query(EventRequest).filter(EventRequest.user_id == current_user.id).all()
    return requests

@app.post("/api/generate-event-description", response_model=DescriptionGenerateResponse)
def generate_event_description(
    request: DescriptionGenerateRequest,
    current_user: User = Depends(get_current_user)
):
    """
    Generate event description based on keywords using AI or template system.
    Admin only endpoint.
    """
    print(f"🔔 Generate description request received from user: {current_user.email}, role: {current_user.role}")
    print(f"📝 Request data: keywords={request.keywords}, title={request.title}, type={request.type}")
    
    if current_user.role != "admin":
        print(f"❌ Access denied: user {current_user.email} is not admin")
        raise HTTPException(status_code=403, detail="Тек администраторлар сипаттамаларды генерациялай алады")
    
    keywords = request.keywords.lower()
    title = request.title or ""
    event_type = request.type or "workshop"
    audience = request.audience or "students"
    
    print(f"✅ Processing description generation for: {keywords}")
    
    # Import random for variety in both AI and template systems
    import random
    import re
    
    # Try to use OpenAI API if available
    openai_api_key = os.getenv("OPENAI_API_KEY")
    if openai_api_key:
        try:
            from openai import OpenAI
            # Initialize client without proxies parameter to avoid compatibility issues
            client = OpenAI(
                api_key=openai_api_key,
                timeout=60.0  # Increase timeout for longer responses
            )
            
            # Build enhanced prompt for AI with more context and variety
            
            # Add variety to system message with emphasis on uniqueness
            import time
            timestamp = int(time.time())
            system_messages = [
                f"Сіз колледж іс-шаралары үшін кәсіби, әртүрлі және қызықты сипаттамалар жасайтын көмекшісіз. Қазақ тілінде жазыңыз. ӘРБІР СИПАТТАМА МУТЛЯҚ БІРЕГЕЙ БОЛУЫ КЕРЕК - басқа сипаттамаларға ұқсамауы керек. Әрбір сипаттаманы бірегей, креативті және тартымды етіп жасаңыз. Шаблондарды қолданбаңыз!",
                f"Сіз студенттерге арналған іс-шаралардың сипаттамаларын жасайтын мамансыз. Қазақ тілінде, әртүрлі стильде, қызықты және мотивациялық сипаттамалар жасаңыз. ӘРБІР СИПАТТАМА БІРЕГЕЙ БОЛУЫ КЕРЕК - басқа сипаттамаларға ұқсамауы керек. Шаблондарды қайталамаңыз!",
                f"Сіз креативті және кәсіби сипаттамалар жасайтын көмекшісіз. Қазақ тілінде, әрбір сипаттаманы басқасынан МУТЛЯҚ ЕРЕКШЕЛЕП, студенттерді қызықтыратын етіп жазыңыз. ӘРБІР СИПАТТАМА БІРЕГЕЙ БОЛУЫ КЕРЕК - шаблондарды қолданбаңыз!"
            ]
            
            system_message = random.choice(system_messages)
            
            # Enhanced prompt with more context - REQUIRES MINIMUM 250 WORDS AND UNIQUENESS
            # Add timestamp and random seed for uniqueness
            unique_seed = f"{timestamp}_{random.randint(1000, 9999)}"
            prompt = f"""Сіз колледж іс-шаралары үшін ДЕТАЛЬДЫ, КРЕАТИВТІ, МОТИВАЦИЯЛЫҚ, ТОЛЫҚ және МУТЛЯҚ БІРЕГЕЙ сипаттамалар жасайтын мамансыз.

⚠️ КРИТИКАЛЫҚ ТӘЛІМДЕР (МУТЛЯҚ ОРЫНДАЛУЫ КЕРЕК):
• МИНИМУМ 250 СӨЗ (қазақ тілінде) - бұл міндетті талап, орындалуы керек!
• ӘРБІР СИПАТТАМА МУТЛЯҚ БІРЕГЕЙ БОЛУЫ КЕРЕК - басқа сипаттамаларға ұқсамауы керек!
• Шаблондарды қолданбаңыз - әрбір сипаттама бірегей болуы керек!
• Әртүрлі сөздер, фразалар, идеялар мен детальдар қолданыңыз!
• Креативті ойлаңыз - стандартты фразаларды қолданбаңыз! 

⚠️ БІРЕГЕЙЛІК ТӘЛІМДЕРІ:
• ӘРБІР СИПАТТАМА МУТЛЯҚ БІРЕГЕЙ БОЛУЫ КЕРЕК - басқа сипаттамаларға ұқсамауы керек!
• Стандартты фразаларды қолданбаңыз: "Бұл іс-шара...", "Студенттер...", "Іс-шарада..." сияқты бастауларды әртүрлі етіп жазыңыз!
• Әртүрлі сөздер мен фразалар қолданыңыз - қайталамаңыз!
• Креативті бастаулар қолданыңыз - әрбір сипаттама басқасынан ерекше басталуы керек!
• Детальдарды әртүрлі түрде сипаттаңыз - шаблондық тіл қолданбаңыз!

═══════════════════════════════════════════════════════════════
КІЛТ АҚПАРАТ (мұқият талдаңыз және БАРЛЫҒЫН пайдаланыңыз):
═══════════════════════════════════════════════════════════════
• Кілт сөздер: "{request.keywords}"
• Атауы: {title if title else 'көрсетілмеген'}
• Іс-шара типі: {event_type}
• Аудитория: {audience}
• Күні: {request.date if request.date else 'көрсетілмеген'}
• Орын: {request.location if request.location else 'көрсетілмеген'}

⚠️ ЕСКЕРТУ: Кілт сөздерде адамдардың аттары, топтар, ерекше детальдар болса - оларды МУТЛЯҚ пайдаланыңыз!

═══════════════════════════════════════════════════════════════
ТАЛДАУ ЖӘНЕ КРЕАТИВТІ ОЙЛАУ:
═══════════════════════════════════════════════════════════════
1. Кілт сөздерді терең талдаңыз:
   - Бұл қандай іс-шара? (концерт, семинар, воркшоп, жарыс, кездесу, т.б.)
   - Қандай эмоциялар мен күтімдер бар?
   - Қандай ерекшеліктер мен детальдар көрсетілген?

2. Контекстті анықтаңыз:
   - Егер "концерт", "лotoрея", "топ", "группа", "көрініс", "мереке" болса → мәдени-көркем іс-шара
   - Егер "воркшоп", "семинар", "лекция", "білім", "тәжірибе" болса → білім беру іс-шарасы
   - Егер "жарыс", "турнир", "соревнование" болса → спорттық іс-шара

3. Детальдарды табыңыз және пайдаланыңыз:
   - Топтар (2F1, 2F2, 1A1, т.б.) - МУТЛЯҚ пайдаланыңыз!
   - Адамдардың аттары - егер көрсетілген болса, оларды пайдаланыңыз!
   - Ерекше элементтер (лотерея, сыйлықтар, қонақтар, т.б.)
   - Тақырыптар мен бағыттар
   - Күні, уақыты, орыны - егер көрсетілген болса, пайдаланыңыз!

═══════════════════════════════════════════════════════════════
СИПАТТАМА ҚҰРЫЛЫМЫ (МИНИМУМ 250 СӨЗ, детальды, креативті, толық):
═══════════════════════════════════════════════════════════════

1. ШОЛУ (8-12 сөйлем, ДЕТАЛЬДЫ, ТАРТЫМДЫ және МУТЛЯҚ БІРЕГЕЙ):
   ⚠️ МАҢЫЗДЫ: Бұл секция МИНИМУМ 8-12 сөйлем болуы керек (250+ сөз үшін)!
   ⚠️ БІРЕГЕЙЛІК: Стандартты "Бұл іс-шара..." бастауын қолданбаңыз! Креативті бастаулар қолданыңыз!
   - Кілт сөздердегі БАРЛЫҚ ақпаратты пайдаланыңыз (топтар, адамдардың аттары, ерекше детальдар)
   - Іс-шараның мақсаты мен маңызын ДЕТАЛЬДЫ және БІРЕГЕЙ түрде сипаттаңыз
   - Неге бұл іс-шара қызықты және маңызды екенін ТЕРЕҢ, ДЕТАЛЬДЫ және БІРЕГЕЙ түсіндіріңіз
   - Егер концерт болса: мерекелік атмосфера, таланттар, көріністер, лотерея, топтар, адамдардың аттары сияқты ДЕТАЛЬДАРДЫ БІРЕГЕЙ түрде қосыңыз
   - Егер воркшоп болса: практикалық мән, білім, тәжірибе, мүмкіндіктер, тақырыптар сияқты аспектілерді ДЕТАЛЬДЫ және БІРЕГЕЙ көрсетіңіз
   - Егер адамдардың аттары көрсетілген болса - оларды МУТЛЯҚ пайдаланыңыз!
   - Егер топтар көрсетілген болса - оларды МУТЛЯҚ пайдаланыңыз!
   - ӘРТҮРЛІ бастаулар қолданыңыз - "Бұл іс-шара...", "Студенттер...", "Іс-шарада..." сияқты стандартты фразаларды қолданбаңыз!
   - Креативті, тартымды, детальды және БІРЕГЕЙ тіл қолданыңыз
   - Көптеген идеялар, детальдар және БІРЕГЕЙ элементтер қосыңыз
   - Әрбір сөйлемді әртүрлі, креативті және детальды етіп жазыңыз

2. КІМ ҚАТЫСУҒА БОЛАДЫ (7-10 пункт, ДЕТАЛЬДЫ және БІРЕГЕЙ):
   ⚠️ МАҢЫЗДЫ: Бұл секция МИНИМУМ 7-10 пункт болуы керек (250+ сөз үшін)!
   ⚠️ БІРЕГЕЙЛІК: Стандартты "• Бұл іс-шара..." формулировкасын қолданбаңыз! Әртүрлі бастаулар қолданыңыз!
   - Кілт сөздерде көрсетілген топтарды МУТЛЯҚ нақты және БІРЕГЕЙ түрде көрсетіңіз (егер бар болса)
   - Егер адамдардың аттары көрсетілген болса - оларды пайдаланыңыз!
   - ӘРТҮРЛІ формулировкалар қолданыңыз - шаблондарды қайталамаңыз!
   - Аудиторияның ерекшеліктерін ДЕТАЛЬДЫ және БІРЕГЕЙ сипаттаңыз
   - Қатысуға қызығушылық танытушыларды ДЕТАЛЬДЫ, креативті және БІРЕГЕЙ сипаттаңыз
   - Көптеген идеялар, детальдар және БІРЕГЕЙ элементтер қосыңыз
   - Әрбір пунктті әртүрлі, детальды және креативті етіп жазыңыз
   - Мысалдар (БІРЕГЕЙ формулировкалар): "[Топтар] студенттеріне арналған ерекше мүмкіндік", "[Топтар] студенттерінің қатысуына ерекше назар аударылады", "Негізінен [топтар] студенттеріне бағытталған, бірақ барлығы қатысуға шақырылады", "[Топтар] студенттерінің қызығушылығына сәйкес келетін іс-шара", т.б.

3. ІС-ШАРА МАЗМҰНЫ / НЕ ҮЙРЕНЕДІ (8-12 пункт, ДЕТАЛЬДЫ, креативті және БІРЕГЕЙ):
   ⚠️ МАҢЫЗДЫ: Бұл секция МИНИМУМ 8-12 пункт болуы керек (250+ сөз үшін)!
   ⚠️ БІРЕГЕЙЛІК: Стандартты формулировкаларды қолданбаңыз! Әрбір пунктті әртүрлі, детальды және креативті етіп жазыңыз!
   - Егер концерт/көрініс болса: "Іс-шара мазмұны:" деп бастаңыз
     * Нақты мазмұнды ДЕТАЛЬДЫ сипаттаңыз (концерт, лотерея, көріністер, нөмірлер, топтар, адамдардың аттары, т.б.)
     * Әр элементті айқын, қызықты және детальды етіп сипаттаңыз
     * "Үйрену", "білім", "воркшоп" сияқты сөздерді қолданбаңыз
     * Эмоциялар мен атмосфераны ДЕТАЛЬДЫ сипаттаңыз
     * Кілт сөздердегі БАРЛЫҚ детальдарды пайдаланыңыз
     * Көптеген идеялар мен креативті элементтер қосыңыз
   
   - Егер воркшоп/семинар болса: "Қатысушылар не үйренеді:" деп бастаңыз
     * Практикалық дағдылар мен білімдерді ДЕТАЛЬДЫ сипаттаңыз
     * Конкретті мүмкіндіктер мен нәтижелерді ДЕТАЛЬДЫ көрсетіңіз
     * Тәжірибе мен практика аспектілерін ерекшелеңіз
     * Білім беру мақсатын айқындаңыз
     * Кілт сөздердегі тақырыптар мен бағыттарды пайдаланыңыз
     * Көптеген идеялар мен практикалық мысалдар қосыңыз
   
   - Әр пунктті әртүрлі, детальды, креативті және нақты етіп жазыңыз
   - Шаблондық фразаларды қолданбаңыз, креативті ойлаңыз
   - Кілт сөздердегі БАРЛЫҚ ақпаратты пайдаланыңыз

4. ЕРЕКШЕ ДЕТАЛЬДАР (5-8 пункт, ДЕТАЛЬДЫ және БІРЕГЕЙ):
   ⚠️ МАҢЫЗДЫ: Бұл секция МИНИМУМ 5-8 пункт болуы керек (250+ сөз үшін)!
   ⚠️ БІРЕГЕЙЛІК: Әрбір пунктті әртүрлі, детальды және креативті етіп жазыңыз!
   - Лотерея, сыйлықтар, қонақтар, ерекше бағдарламалар сияқты детальдарды ДЕТАЛЬДЫ және БІРЕГЕЙ түрде қосыңыз
   - Егер кілт сөздерде ерекше элементтер көрсетілген болса, оларды ДЕТАЛЬДЫ және БІРЕГЕЙ сипаттаңыз
   - Егер адамдардың аттары көрсетілген болса - оларды пайдаланыңыз!
   - Егер топтар көрсетілген болса - оларды пайдаланыңыз!
   - Көптеген идеялар, креативті элементтер және БІРЕГЕЙ детальдар қосыңыз
   - Әрбір пунктті әртүрлі, детальды және креативті етіп жазыңыз

5. ҚОСЫМША АҚПАРАТ (4-6 пункт, креативті және БІРЕГЕЙ):
   ⚠️ МАҢЫЗДЫ: Бұл секция МИНИМУМ 4-6 пункт болуы керек (250+ сөз үшін)!
   ⚠️ БІРЕГЕЙЛІК: Әрбір пунктті әртүрлі, детальды және креативті етіп жазыңыз!
   - Іс-шараның маңызы мен құндылығын БІРЕГЕЙ және детальды сипаттаңыз
   - Қатысушыларға не күту керектігін детальды, креативті және БІРЕГЕЙ түсіндіріңіз
   - Мотивациялық элементтерді БІРЕГЕЙ және креативті қосыңыз
   - Іс-шараның ерекшеліктерін және артықшылықтарын детальды сипаттаңыз
   - Қатысушыларға арналған қосымша мүмкіндіктер мен артықшылықтарды сипаттаңыз

═══════════════════════════════════════════════════════════════
ЕРЕЖЕЛЕР ЖӘНЕ ТӘЛІМДЕР:
═══════════════════════════════════════════════════════════════
✅ МИНИМУМ 250 СӨЗ (қазақ тілінде) - бұл міндетті талап, орындалуы керек! Егер 250 сөзден аз болса - қайта жазыңыз!
✅ Кілт сөздердегі БАРЛЫҚ ақпаратты дәл қолданыңыз (топтар, адамдардың аттары, ерекше детальдар)
✅ Егер адамдардың аттары көрсетілген болса - оларды МУТЛЯҚ пайдаланыңыз!
✅ Егер топтар көрсетілген болса - оларды МУТЛЯҚ пайдаланыңыз!
✅ ӘРБІР СИПАТТАМА МУТЛЯҚ БІРЕГЕЙ БОЛУЫ КЕРЕК - басқа сипаттамаларға ұқсамауы керек!
✅ Шаблондарды қайталамаңыз - әрбір сипаттама бірегей болуы керек!
✅ Стандартты фразаларды қолданбаңыз - "Бұл іс-шара...", "Студенттер...", "Іс-шарада..." сияқты бастауларды әртүрлі етіп жазыңыз!
✅ Әртүрлі сөздер, фразалар, идеялар мен детальдар қолданыңыз - қайталамаңыз!
✅ Егер концерт болса - "үйрену", "білім", "воркшоп" сияқты сөздерді қолданбаңыз
✅ Егер воркшоп болса - білім беру аспектісін детальды көрсетіңіз
✅ Қарапайым, түсінікті, бірақ қызықты, тартымды және детальды тіл
✅ Кәсіби, мотивациялық, бірақ табиғи және жылы тон
✅ Эмодзи қолданбаңыз
✅ Маркетингтік преувелификация жоқ, бірақ мотивациялық болуы керек
✅ Колледж ортасына сәйкес, студенттерге арналған
✅ Тек қазақ тілінде
✅ Детальды, толық және анықтамалық сипаттама - қысқаша емес!
✅ Креативті ойлаңыз - шаблондарды қолданбаңыз
✅ Әрбір сипаттамада әртүрлі сөздер, фразалар, идеялар және детальдар қолданыңыз
✅ Көптеген идеялар мен креативті элементтер қосыңыз

═══════════════════════════════════════════════════════════════
МЫСАЛ (концерт үшін - детальды және креативті):
═══════════════════════════════════════════════════════════════
"Бұл мерекелік концерт студенттердің шығармашылық қабілеттерін көрсетуге арналған ерекше іс-шара. Іс-шарада талантты студенттер өз өнерлерін көрсетеді, әртүрлі музыкалық және би нөмірлерімен келушілерді таң қалдырады. Сондай-ақ, қызықты лотерея өткізіліп, бақыттылар сыйлықтарға ие болады. Бұл іс-шара мерекелік атмосферада демалу және достарымен бірге уақыт өткізу үшін тамаша мүмкіндік.

Кім қатысуға болады:
• 2F1, 2F2 топтарының студенттері - бұл іс-шара негізінен оларға арналған
• Көркем нөмірлер мен көріністерді көруге қызығушылық танытқан барлық студенттер
• Достарымен бірге уақыт өткізгісі келетіндер
• Мәдени-көркем іс-шараларға қатысқысы келетіндер
• Мерекелік көңіл-күй іздегендер

Іс-шара мазмұны:
• Студенттердің музыкалық және би нөмірлері - әртүрлі стильдер мен жанрлар
• Қызықты лотерея ойыны және сыйлықтар тарату
• Мерекелік атмосферада демалу мүмкіндігі
• Талантты студенттердің өнерін тамашалау
• Достарымен бірге уақыт өткізу және жаңа танысулар"

═══════════════════════════════════════════════════════════════
МАҢЫЗДЫ ТӘЛІМДЕР:
• МИНИМУМ 250 СӨЗ (қазақ тілінде) - бұл міндетті талап!
• Кілт сөздердегі БАРЛЫҚ ақпаратты пайдаланыңыз (топтар, адамдардың аттары, ерекше детальдар)
• Егер адамдардың аттары көрсетілген болса - оларды МУТЛЯҚ пайдаланыңыз!
• Егер топтар көрсетілген болса - оларды МУТЛЯҚ пайдаланыңыз!
• Креативті ойлаңыз - шаблондарды қолданбаңыз
• Көптеген идеялар мен детальдар қосыңыз
• Детальды, толық және анықтамалық сипаттама жасаңыз

═══════════════════════════════════════════════════════════════
ФИНАЛЬНЫЕ ТӘЛІМДЕР:
═══════════════════════════════════════════════════════════════
• МИНИМУМ 250 СӨЗ (қазақ тілінде) - бұл міндетті талап, орындалуы керек!
• ӘРБІР СИПАТТАМА МУТЛЯҚ БІРЕГЕЙ БОЛУЫ КЕРЕК - басқа сипаттамаларға ұқсамауы керек!
• Стандартты фразаларды қолданбаңыз - "Бұл іс-шара...", "Студенттер...", "Іс-шарада..." сияқты бастауларды әртүрлі етіп жазыңыз!
• Әртүрлі сөздер, фразалар, идеялар мен детальдар қолданыңыз - қайталамаңыз!
• Креативті ойлаңыз - шаблондарды қолданбаңыз!
• Көптеген идеялар, детальдар және БІРЕГЕЙ элементтер қосыңыз!
• Детальды, толық, анықтамалық және МУТЛЯҚ БІРЕГЕЙ сипаттама жасаңыз!

Енді кілт сөздерге сәйкес, бірақ МИНИМУМ 250 СӨЗ (қазақ тілінде), детальды, креативті, толық және МУТЛЯҚ БІРЕГЕЙ сипаттама жасаңыз. Кілт сөздердегі БАРЛЫҚ ақпаратты (топтар, адамдардың аттары, ерекше детальдар) пайдаланыңыз! Шаблондарды қолданбаңыз, креативті ойлаңыз, көптеген идеялар қосыңыз және әрбір сипаттаманы басқасынан МУТЛЯҚ ЕРЕКШЕЛЕП жазыңыз!"""
            
            # Use gpt-4o-mini for better quality and longer responses
            response = client.chat.completions.create(
                model="gpt-4o-mini",  # Better model for longer, more creative descriptions
                messages=[
                    {"role": "system", "content": system_message},
                    {"role": "user", "content": prompt}
                ],
                temperature=1.0,  # Maximum temperature for maximum creativity, uniqueness and variety
                max_tokens=2500,  # Increased significantly for much longer, more detailed descriptions (250+ words = ~2000-2500 tokens)
                top_p=0.99,  # Very high nucleus sampling for maximum diversity, creativity and uniqueness
                frequency_penalty=0.8,  # Very high penalty to reduce repetition and encourage maximum variety and uniqueness
                presence_penalty=0.8  # Very high penalty to encourage diverse topics, ideas and uniqueness
            )
            
            ai_description = response.choices[0].message.content.strip()
            word_count = len(ai_description.split())
            print(f"✅ AI description generated successfully, length: {len(ai_description)} characters, words: {word_count}")
            
            # Verify minimum word count (250 words minimum - MANDATORY)
            if word_count < 250:
                print(f"⚠️ ERROR: Description has only {word_count} words, but MINIMUM 250 WORDS IS MANDATORY. Regenerating with strong emphasis on length and uniqueness...")
                # Try once more with stronger emphasis on length and uniqueness
                retry_prompt = prompt + f"\n\n⚠️ КРИТИКАЛЫҚ ҚАТЕ: Алдыңғы сипаттама тым қысқа болды ({word_count} сөз). МИНИМУМ 250 СӨЗ (қазақ тілінде) - бұл міндетті талап, орындалуы керек! Енді МИНИМУМ 250 СӨЗ (қазақ тілінде) жазыңыз! Детальды, толық, анықтамалық, креативті және МУТЛЯҚ БІРЕГЕЙ сипаттама жасаңыз! Көптеген детальдар, идеялар, ақпарат және БІРЕГЕЙ элементтер қосыңыз! Әрбір секцияны кеңейтіңіз: ШОЛУ (12+ сөйлем), КІМ ҚАТЫСУҒА БОЛАДЫ (10+ пункт), ІС-ШАРА МАЗМҰНЫ (12+ пункт), ЕРЕКШЕ ДЕТАЛЬДАР (8+ пункт), ҚОСЫМША АҚПАРАТ (6+ пункт)!"
                retry_response = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        {"role": "system", "content": system_message},
                        {"role": "user", "content": retry_prompt}
                    ],
                    temperature=1.0,  # Maximum temperature for maximum creativity and uniqueness
                    max_tokens=3000,  # Even more tokens for retry to ensure 250+ words
                    top_p=0.99,
                    frequency_penalty=0.9,  # Very high penalty for maximum uniqueness
                    presence_penalty=0.9  # Very high penalty for maximum uniqueness
                )
                ai_description = retry_response.choices[0].message.content.strip()
                word_count = len(ai_description.split())
                print(f"✅ Retry: Description length: {len(ai_description)} characters, words: {word_count}")
                
                # If still less than 250 words, try one more time
                if word_count < 250:
                    print(f"⚠️ ERROR: Retry still has only {word_count} words. Final attempt with maximum emphasis...")
                    final_prompt = f"КІЛТ СӨЗДЕР: {request.keywords}\nАТАУЫ: {title}\nТИПІ: {event_type}\n\n⚠️ КРИТИКАЛЫҚ: МИНИМУМ 250 СӨЗ (қазақ тілінде) жазыңыз! Бұл міндетті талап! Детальды, толық, креативті және МУТЛЯҚ БІРЕГЕЙ сипаттама жасаңыз! Көптеген детальдар, идеялар және ақпарат қосыңыз! Әрбір секцияны кеңейтіңіз!"
                    final_response = client.chat.completions.create(
                        model="gpt-4o-mini",
                        messages=[
                            {"role": "system", "content": "Сіз колледж іс-шаралары үшін МИНИМУМ 250 СӨЗ (қазақ тілінде) детальды, креативті және МУТЛЯҚ БІРЕГЕЙ сипаттамалар жасайтын мамансыз. Қазақ тілінде жазыңыз."},
                            {"role": "user", "content": final_prompt}
                        ],
                        temperature=1.0,
                        max_tokens=3500,
                        top_p=0.99,
                        frequency_penalty=0.9,
                        presence_penalty=0.9
                    )
                    ai_description = final_response.choices[0].message.content.strip()
                    word_count = len(ai_description.split())
                    print(f"✅ Final attempt: Description length: {len(ai_description)} characters, words: {word_count}")
            
            return DescriptionGenerateResponse(description=ai_description)
            
        except Exception as e:
            print(f"❌ OpenAI API error: {e}")
            print(f"📋 Falling back to template system")
            # Fall back to template system if AI fails
            pass
    
    # Event type translations
    type_map = {
        "workshop": "воркшоп",
        "seminar": "семинар",
        "competition": "жарыс",
        "meetup": "кездесу",
        "concert": "концерт",
        "event": "іс-шара"
    }
    type_kz = type_map.get(event_type, "іс-шара")
    
    # Audience translations
    audience_map = {
        "beginner": "бастапқы деңгейдегі",
        "advanced": "тәжірибелі",
        "students": "студенттер"
    }
    audience_kz = audience_map.get(audience, "студенттер")
    
    # Extract main topic from keywords
    topics = []
    if "ai" in keywords or "жасалма интеллект" in keywords or "искусственный интеллект" in keywords:
        topics.append("жасалма интеллект")
    if "programming" in keywords or "программирование" in keywords or "бағдарламалау" in keywords:
        topics.append("бағдарламалау")
    if "design" in keywords or "дизайн" in keywords:
        topics.append("дизайн")
    if "marketing" in keywords or "маркетинг" in keywords:
        topics.append("маркетинг")
    if "business" in keywords or "бизнес" in keywords:
        topics.append("бизнес")
    if "sport" in keywords or "спорт" in keywords:
        topics.append("спорт")
    if "music" in keywords or "музыка" in keywords or "ән" in keywords:
        topics.append("ән")
    if "art" in keywords or "сурет" in keywords or "көркем" in keywords:
        topics.append("көркем өнер")
    
    topic_text = ", ".join(topics) if topics else "тақырып"
    
    # Generate description based on template
    description_parts = []
    
    # Check if it's a concert or entertainment event
    is_concert = "концерт" in keywords or "concert" in keywords or event_type == "concert"
    is_entertainment = "лотерея" in keywords or "lottery" in keywords or "көрініс" in keywords or "лotoрея" in keywords
    
    # Overview (5-8 sentences for 250+ words) - with variety and more detail
    if is_concert or is_entertainment:
        # For concerts and entertainment events - multiple variations
        concert_intros = [
            f"Бұл мерекелік {type_kz} студенттердің таланттарын көрсетуге арналған.",
            f"Бұл {type_kz} студенттердің шығармашылық қабілеттерін көрсетуге арналған.",
            f"Бұл {type_kz} көркем нөмірлер мен көріністерді көруге мүмкіндік береді.",
            f"Бұл {type_kz} студенттердің өнерін тамашалауға арналған."
        ]
        
        if title:
            title_variations = [
                f"Бұл мерекелік {type_kz} «{title}» тақырыбына арналған ерекше іс-шара. Бұл іс-шара студенттердің шығармашылық қабілеттерін көрсетуге арналған және колледж қауымдастығының бірлігін нығайтуға бағытталған. Іс-шарада талантты студенттер өз өнерлерін көрсетеді, әртүрлі музыкалық және би нөмірлерімен келушілерді таң қалдырады. Бұл іс-шара мерекелік атмосферада демалу және достарымен бірге уақыт өткізу үшін тамаша мүмкіндік.",
                f"«{title}» тақырыбындағы {type_kz} студенттерді күтеді. Бұл ерекше іс-шара студенттердің таланттарын көрсетуге арналған және колледж қауымдастығының мәдени өмірін байытуға бағытталған. Іс-шарада көркем нөмірлер мен көріністер көрсетіледі, студенттер өз шығармашылық қабілеттерін көрсету мүмкіндігіне ие болады. Бұл іс-шара мерекелік көңіл-күй мен қуанышты бірге бөлісуге мүмкіндік береді.",
                f"Бұл {type_kz} {title} тақырыбын қамтиды және студенттердің шығармашылық қабілеттерін көрсетуге арналған. Іс-шарада талантты студенттер өз өнерлерін көрсетеді, әртүрлі музыкалық және би нөмірлерімен келушілерді таң қалдырады. Бұл іс-шара мерекелік атмосферада демалу және достарымен бірге уақыт өткізу үшін тамаша мүмкіндік. Колледж қауымдастығының бірлігін нығайтуға бағытталған бұл іс-шара барлық қатысушылар үшін ұмытылмас тәжірибе болады."
            ]
            description_parts.append(random.choice(title_variations))
        else:
            extended_intros = [
                f"Бұл мерекелік {type_kz} студенттердің таланттарын көрсетуге арналған ерекше іс-шара. Бұл іс-шара колледж қауымдастығының бірлігін нығайтуға бағытталған және студенттердің шығармашылық қабілеттерін дамытуға мүмкіндік береді. Іс-шарада талантты студенттер өз өнерлерін көрсетеді, әртүрлі музыкалық және би нөмірлерімен келушілерді таң қалдырады. Бұл іс-шара мерекелік атмосферада демалу және достарымен бірге уақыт өткізу үшін тамаша мүмкіндік.",
                f"Бұл {type_kz} студенттердің шығармашылық қабілеттерін көрсетуге арналған ерекше іс-шара. Бұл іс-шара колледж қауымдастығының мәдени өмірін байытуға бағытталған және студенттердің таланттарын дамытуға мүмкіндік береді. Іс-шарада көркем нөмірлер мен көріністер көрсетіледі, студенттер өз шығармашылық қабілеттерін көрсету мүмкіндігіне ие болады. Бұл іс-шара мерекелік көңіл-күй мен қуанышты бірге бөлісуге мүмкіндік береді."
            ]
            description_parts.append(random.choice(extended_intros))
        
        if is_entertainment:
            lottery_variations = [
                "Іс-шара кезінде қызықты лотерея өткізіледі, қатысушылар сыйлықтарға ие бола алады. Лотерея ойыны бағдарламаның ерекше бөлігі болып табылады және барлық қатысушыларға тең мүмкіндік береді. Бақыттылар әртүрлі сыйлықтарға ие болады, бұл іс-шараны одан да қызықты етеді. Лотерея кезінде мерекелік атмосфера мен қуаныш сақталады, барлық қатысушылар үшін ұмытылмас тәжірибе болады.",
                "Лотерея ойыны мен сыйлықтар тарату бағдарламасына енгізілген. Бұл элемент іс-шараны одан да қызықты етеді және барлық қатысушыларға тең мүмкіндік береді. Лотерея кезінде мерекелік атмосфера мен қуаныш сақталады, бақыттылар әртүрлі сыйлықтарға ие болады. Бұл іс-шараның ерекше бөлігі болып табылады және барлық қатысушылар үшін ұмытылмас тәжірибе болады."
            ]
            description_parts.append(random.choice(lottery_variations))
        else:
            concert_details = [
                "Іс-шарада студенттер өз таланттарын көрсетеді, көріністер мен музыкалық нөмірлер болады. Бұл іс-шара студенттердің шығармашылық қабілеттерін дамытуға мүмкіндік береді және колледж қауымдастығының мәдени өмірін байытуға бағытталған. Іс-шарада әртүрлі стильдер мен жанрлар көрсетіледі, студенттер өз өнерлерін көрсету мүмкіндігіне ие болады. Бұл іс-шара мерекелік атмосферада демалу және достарымен бірге уақыт өткізу үшін тамаша мүмкіндік.",
                "Студенттердің көркем нөмірлері мен көріністері көрсетіледі. Бұл іс-шара студенттердің таланттарын көрсетуге арналған және колледж қауымдастығының бірлігін нығайтуға бағытталған. Іс-шарада әртүрлі музыкалық және би нөмірлерімен келушілерді таң қалдыратын талантты студенттер өз өнерлерін көрсетеді. Бұл іс-шара мерекелік көңіл-күй мен қуанышты бірге бөлісуге мүмкіндік береді."
            ]
            description_parts.append(random.choice(concert_details))
    else:
        # For educational events (workshops, seminars) - multiple variations
        if title:
            workshop_intros = [
                f"Бұл {type_kz} «{title}» тақырыбына арналған. {audience_kz.capitalize()} студенттерге арналған бұл іс-шарада сіз практикалық дағдылар мен білім ала аласыз.",
                f"«{title}» тақырыбындағы {type_kz} {audience_kz} студенттерге арналған. Бұл іс-шарада практикалық тәжірибе мен білім беріледі.",
                f"Бұл {type_kz} {title} тақырыбын қамтиды. {audience_kz.capitalize()} студенттер үшін практикалық дағдылар дамытуға бағытталған."
            ]
            description_parts.append(random.choice(workshop_intros))
        else:
            workshop_generic = [
                f"Бұл {type_kz} {topic_text} бағытындағы білім мен тәжірибені бөлісуге арналған. {audience_kz.capitalize()} студенттерге арналған бұл іс-шарада сіз практикалық дағдылар дамыта аласыз.",
                f"Бұл {type_kz} {topic_text} тақырыбына арналған. {audience_kz.capitalize()} студенттер үшін практикалық білім беру мақсатында ұйымдастырылған.",
                f"{topic_text.capitalize()} бағытындағы {type_kz} {audience_kz} студенттерге арналған. Бұл іс-шарада практикалық тәжірибе мен білім беріледі."
            ]
            description_parts.append(random.choice(workshop_generic))
        
        workshop_details = [
            "Іс-шарада қазіргі заманғы әдістер мен технологиялармен танысасыз, сұрақтар қоя аласыз және тәжірибелі мамандармен байланыс орната аласыз.",
            "Іс-шара кезінде практикалық тапсырмалар орындайсыз, жаңа әдістер үйреніп, тәжірибелі мамандармен білім алмасасыз.",
            "Іс-шарада қазіргі заманғы трендтер мен технологиялармен танысасыз, практикалық дағдылар дамытасыз және сұрақтар қоя аласыз."
        ]
        description_parts.append(random.choice(workshop_details))
    
    # Who should attend (expanded for 250+ words)
    description_parts.append(f"\nКім қатысуға болады:")
    
    # Extract groups from keywords
    groups = []
    group_pattern = r'\b(\d+[A-Z]\d+)\b'
    found_groups = re.findall(group_pattern, request.keywords, re.IGNORECASE)
    if found_groups:
        groups = found_groups
    
    if is_concert or is_entertainment:
        if groups:
            group_variations = [
                f"• {', '.join(groups)} топтарының студенттері - бұл іс-шара негізінен оларға арналған және олардың қатысуына ерекше назар аударылады",
                f"• Негізінен {', '.join(groups)} топтарының студенттеріне арналған, бірақ барлық студенттер қатысуға шақырылады",
                f"• {', '.join(groups)} топтары студенттері қатысуға шақырылады және олардың қатысуына ерекше назар аударылады"
            ]
            description_parts.append(random.choice(group_variations))
        else:
            all_students = [
                "• Барлық студенттер - бұл іс-шара колледжтің барлық студенттеріне ашық және барлығы қатысуға шақырылады",
                "• Колледжтің барлық студенттері - бұл іс-шара барлық курс студенттеріне арналған және олардың қатысуына қызығушылық танытады",
                "• Барлық курс студенттері - бұл іс-шара барлық деңгейдегі студенттерге арналған және олардың қатысуына мүмкіндік береді"
            ]
            description_parts.append(random.choice(all_students))
        
        audience_variations = [
            "• Көркем нөмірлер мен көріністерді көруге қызығушылық танытқандар - бұл іс-шара оларға тамаша мүмкіндік береді",
            "• Мәдени-көркем іс-шараларға қатысқысы келетіндер - бұл іс-шара олардың қызығушылығына сәйкес келеді",
            "• Өнер мен мәдениетті құрметтейтіндер - бұл іс-шара оларға өнерді тамашалау мүмкіндігін береді",
            "• Достарымен бірге уақыт өткізгісі келетіндер - бұл іс-шара оларға тамаша мүмкіндік береді",
            "• Мерекелік көңіл-күй іздегендер - бұл іс-шара оларға қуаныш пен мерекелік атмосфера береді",
            "• Шығармашылық қабілеттерін көрсеткісі келетіндер - бұл іс-шара оларға өз таланттарын көрсету мүмкіндігін береді"
        ]
        # Add more items for longer description
        selected_audience = random.sample(audience_variations, min(5, len(audience_variations)))
        description_parts.extend(selected_audience)
    else:
        if audience == "beginner":
            description_parts.append(f"• Бастапқы деңгейдегі студенттер")
            description_parts.append(f"• Тақырыпты жаңадан үйренгісі келетіндер")
            description_parts.append(f"• Негізгі білім алуға қызығушылық танытқандар")
        elif audience == "advanced":
            description_parts.append(f"• Тәжірибелі студенттер")
            description_parts.append(f"• Білімдерін тереңдеткісі келетіндер")
            description_parts.append(f"• Кәсіби дағдыларды дамытқысы келетіндер")
        else:
            description_parts.append(f"• Барлық деңгейдегі студенттер")
            description_parts.append(f"• Тақырыпқа қызығушылық танытқандар")
            description_parts.append(f"• Жаңа білім мен тәжірибе алуға дайындар")
    
    # What participants will learn or what will happen - with variety
    if is_concert or is_entertainment:
        description_parts.append(f"\nІс-шара мазмұны:")
        if is_entertainment:
            lottery_content = [
                "• Қызықты лотерея ойыны",
                "• Сыйлықтар тарату",
                "• Бақыттыларды анықтау"
            ]
            description_parts.extend(random.sample(lottery_content, 2))
        
        concert_content = [
            "• Студенттердің көркем нөмірлері мен көріністері",
            "• Мәдени-көркем бағдарлама",
            "• Музыкалық және би нөмірлері",
            "• Талантты студенттердің өнер көрсетуі"
        ]
        description_parts.extend(random.sample(concert_content, 2))
    else:
        description_parts.append(f"\nҚатысушылар не үйренеді:")
        
        # Topic-specific learning outcomes with variety
        if "ai" in keywords or "жасалма интеллект" in keywords:
            ai_learnings = [
                ["• Жасалма интеллект технологияларының негіздері", "• AI құралдарын пайдалану әдістері", "• Практикалық мысалдар мен кейстер"],
                ["• AI технологияларының қолданылуы", "• Машиндық оқыту негіздері", "• AI құралдарымен жұмыс істеу"],
                ["• Жасалма интеллекттің қазіргі мүмкіндіктері", "• AI платформаларын танысу", "• Практикалық AI қолданбалары"]
            ]
            description_parts.extend(random.choice(ai_learnings))
        elif "programming" in keywords or "бағдарламалау" in keywords or "python" in keywords or "javascript" in keywords:
            prog_learnings = [
                ["• Бағдарламалау негіздері мен әдістері", "• Практикалық бағдарламалау тапсырмалары", "• Код жазу және оны оптимизациялау"],
                ["• Бағдарламалау тілдерінің негіздері", "• Алгоритмдер мен деректер құрылымдары", "• Практикалық жобалар дамыту"],
                ["• Код жазу дағдыларын дамыту", "• Бағдарламалау әдістері мен практикалары", "• Жобаларды жоспарлау және іске асыру"]
            ]
            description_parts.extend(random.choice(prog_learnings))
        elif "design" in keywords or "дизайн" in keywords:
            design_learnings = [
                ["• Дизайн принциптері мен трендтері", "• Дизайн құралдарын пайдалану", "• Креативті шешімдер табу"],
                ["• Графикалық дизайн негіздері", "• Дизайн құралдары мен бағдарламалары", "• Визуалды коммуникация"],
                ["• Дизайн түрлері мен стильдері", "• Дизайн процесі мен әдістері", "• Портфолио дайындау"]
            ]
            description_parts.extend(random.choice(design_learnings))
        elif "marketing" in keywords or "маркетинг" in keywords:
            marketing_learnings = [
                ["• Маркетинг стратегиялары мен әдістері", "• Диджитал маркетинг құралдары", "• Брендинг және жарнама"],
                ["• Маркетинг негіздері мен принциптері", "• Социальдық медиа маркетинг", "• Маркетинг жоспарлау"],
                ["• Маркетинг каналы мен стратегиялары", "• Контент маркетинг", "• Маркетинг аналитикасы"]
            ]
            description_parts.extend(random.choice(marketing_learnings))
        elif "business" in keywords or "бизнес" in keywords:
            business_learnings = [
                ["• Бизнес негіздері мен стратегиялары", "• Кәсіпкерлік дағдылары", "• Бизнес-жоспарлау және басқару"],
                ["• Бизнес моделдері мен стратегиялары", "• Кәсіпкерлік негіздері", "• Бизнес-жоспарлау әдістері"],
                ["• Бизнес-басқару принциптері", "• Кәсіпкерлік дағдыларын дамыту", "• Бизнес-стратегиялары"]
            ]
            description_parts.extend(random.choice(business_learnings))
        elif "sport" in keywords or "спорт" in keywords or "футбол" in keywords or "баскетбол" in keywords:
            sport_learnings = [
                ["• Спорттық дағдылар мен техникалар", "• Командалық жұмыс және стратегия", "• Денсаулықты сақтау және жаттығу"],
                ["• Спорттық техникалар мен тактикалар", "• Командалық ойын стратегиялары", "• Физикалық дайындық"],
                ["• Спорттық дағдыларды дамыту", "• Командалық жұмыс принциптері", "• Денсаулықты сақтау"]
            ]
            description_parts.extend(random.choice(sport_learnings))
        elif "music" in keywords or "музыка" in keywords or "ән" in keywords:
            music_learnings = [
                ["• Музыкалық дағдылар мен техникалар", "• Ән айту немесе аспапта ойнау", "• Музыкалық шығармашылық"],
                ["• Музыка теориясы мен практикасы", "• Ән айту техникалары", "• Музыкалық шығармашылық"],
                ["• Музыкалық дағдыларды дамыту", "• Аспапта ойнау техникалары", "• Музыкалық шығармашылық"]
            ]
            description_parts.extend(random.choice(music_learnings))
        elif "art" in keywords or "сурет" in keywords or "көркем" in keywords:
            art_learnings = [
                ["• Көркем өнер техникалары", "• Шығармашылық дағдылар", "• Өнертану негіздері"],
                ["• Көркем өнер әдістері мен техникалары", "• Шығармашылық процесі", "• Өнертану принциптері"],
                ["• Көркем өнер техникаларын дамыту", "• Шығармашылық дағдылар", "• Өнертану негіздері"]
            ]
            description_parts.extend(random.choice(art_learnings))
        else:
            generic_learnings = [
                ["• Тақырып бойынша негізгі білімдер", "• Практикалық дағдылар мен әдістер", "• Тәжірибелі мамандармен білім алмасу"],
                ["• Негізгі білімдер мен тәжірибе", "• Практикалық дағдылар дамыту", "• Мамандармен білім алмасу"],
                ["• Тақырып бойынша білім алу", "• Практикалық тәжірибе", "• Тәжірибелі мамандармен байланыс"]
            ]
            description_parts.extend(random.choice(generic_learnings))
    
    # Basic event details (if provided)
    if request.date:
        description_parts.append(f"\nІс-шара күні мен уақыты күнтізбеде көрсетілген.")
    if request.location:
        description_parts.append(f"Орын: {request.location}")
    
    description = "\n".join(description_parts)
    word_count = len(description.split())
    
    # If template description is too short, add more content
    if word_count < 250:
        print(f"⚠️ Template description has only {word_count} words, adding more content...")
        additional_content = [
            "\n\nІс-шараның маңызы:",
            "• Бұл іс-шара студенттердің шығармашылық қабілеттерін дамытуға мүмкіндік береді",
            "• Колледж қауымдастығының бірлігін нығайтуға бағытталған",
            "• Студенттердің мәдени өмірін байытуға көмектеседі",
            "• Достарымен бірге уақыт өткізу және жаңа танысулар мүмкіндігі бар",
            "• Мерекелік атмосферада демалу және қуанышты бірге бөлісуге мүмкіндік береді",
            "• Барлық қатысушылар үшін ұмытылмас тәжірибе болады",
            "• Студенттердің таланттарын көрсетуге арналған ерекше мүмкіндік",
            "• Колледж қауымдастығының мәдени өмірін байытуға бағытталған іс-шара"
        ]
        description += "\n".join(additional_content)
        word_count = len(description.split())
        print(f"✅ Extended template description, now {word_count} words")
    
    print(f"✅ Template description generated, length: {len(description)} characters, words: {word_count}")
    return DescriptionGenerateResponse(description=description)

@app.put("/api/event-requests/{request_id}/status")
def update_event_request_status(
    request_id: int,
    status: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Тек администраторлар өтінім мәртебесін жаңарта алады")
    
    if status not in ["approved", "rejected"]:
        raise HTTPException(status_code=400, detail="Status must be 'approved' or 'rejected'")
    
    event_request = db.query(EventRequest).filter(EventRequest.id == request_id).first()
    if not event_request:
        raise HTTPException(status_code=404, detail="Event request not found")
    
    event_request.status = status
    event_request.reviewed_at = datetime.utcnow()
    event_request.reviewed_by = current_user.id
    
    # If approved, create the event
    if status == "approved":
        event = Event(
            title=event_request.title,
            description=event_request.description,
            date=event_request.date,
            start_time=event_request.start_time,
            location=event_request.location,
            max_participants=event_request.max_participants,
            created_by=current_user.id
        )
        db.add(event)
    
    db.commit()
    db.refresh(event_request)
    return event_request

# Get event participants with groups
@app.get("/api/events/{event_id}/participants")
def get_event_participants(
    event_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Тек администраторлар қатысушыларды көре алады")
    
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Іс-шара табылмады")
    
    registrations = db.query(EventRegistration).filter(EventRegistration.event_id == event_id).all()
    participants = []
    group_counts = {}
    
    for reg in registrations:
        user = db.query(User).filter(User.id == reg.user_id).first()
        if user:
            participants.append({
                "id": user.id,
                "full_name": user.full_name,
                "group": user.group,
                "email": user.email
            })
            
            # Count by group
            if user.group:
                group_counts[user.group] = group_counts.get(user.group, 0) + 1
    
    # Format group counts as "GROUP:COUNT"
    group_info = [f"{group}:{count}" for group, count in sorted(group_counts.items())]
    
    return {
        "event_id": event_id,
        "event_title": event.title,
        "participants": participants,
        "group_counts": group_info,
        "total_participants": len(participants)
    }

# Leaderboard endpoint
@app.get("/api/leaderboard")
def get_leaderboard(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    users = db.query(User).filter(User.role == "student").order_by(User.points.desc()).offset(skip).limit(limit).all()
    
    leaderboard = []
    for idx, user in enumerate(users, start=skip + 1):
        leaderboard.append({
            "rank": idx,
            "id": user.id,
            "full_name": user.full_name,
            "group": user.group,
            "points": user.points or 0,
            "photo_url": user.photo_url
        })
    
    return leaderboard

# Reset points monthly (admin only)
@app.post("/api/leaderboard/reset")
def reset_points(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Тек администраторлар ұпайларды қалпына келтіре алады")
    
    # Reset all student points to 0
    students = db.query(User).filter(User.role == "student").all()
    reset_count = 0
    for student in students:
        student.points = 0
        student.points_reset_date = datetime.utcnow()
        reset_count += 1
    
    db.commit()
    return {"message": f"Points reset for {reset_count} students", "reset_count": reset_count}

# Export event history
@app.get("/api/events/history/export")
def export_event_history(
    format: str = "xlsx",  # "xlsx" or "pdf"
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Тек администраторлар тарихты экспорттай алады")
    
    from fastapi.responses import Response
    import io
    
    # Get all events with participants
    events = db.query(Event).order_by(Event.date.desc(), Event.start_time.desc()).all()
    
    if format == "xlsx":
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Event History"
            
            # Headers
            headers = ["Event ID", "Title", "Date", "Time", "Location", "Max Participants", "Total Participants", "Groups", "Created At"]
            ws.append(headers)
            
            # Style headers
            from openpyxl.styles import Font, PatternFill
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Data rows
            for event in events:
                registrations = db.query(EventRegistration).filter(EventRegistration.event_id == event.id).all()
                participants = []
                group_counts = {}
                
                for reg in registrations:
                    user = db.query(User).filter(User.id == reg.user_id).first()
                    if user:
                        participants.append(user.full_name)
                        if user.group:
                            group_counts[user.group] = group_counts.get(user.group, 0) + 1
                
                group_info = ", ".join([f"{group}:{count}" for group, count in sorted(group_counts.items())]) if group_counts else ""
                participants_str = ", ".join(participants) if participants else ""
                
                ws.append([
                    event.id,
                    event.title,
                    event.date.isoformat(),
                    event.start_time.strftime("%H:%M"),
                    event.location,
                    event.max_participants,
                    len(participants),
                    group_info,
                    event.created_at.isoformat() if event.created_at else ""
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return Response(
                content=output.read(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=event_history.xlsx"}
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl is required for XLSX export. Install it with: pip install openpyxl")
    
    elif format == "pdf":
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=letter)
            elements = []
            
            styles = getSampleStyleSheet()
            title = Paragraph("Event History", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 0.2*inch))
            
            # Prepare data
            data = [["Event ID", "Title", "Date", "Time", "Location", "Participants", "Groups"]]
            
            for event in events:
                registrations = db.query(EventRegistration).filter(EventRegistration.event_id == event.id).all()
                group_counts = {}
                for reg in registrations:
                    user = db.query(User).filter(User.id == reg.user_id).first()
                    if user and user.group:
                        group_counts[user.group] = group_counts.get(user.group, 0) + 1
                
                group_info = ", ".join([f"{group}:{count}" for group, count in sorted(group_counts.items())]) if group_counts else ""
                
                data.append([
                    str(event.id),
                    event.title[:30],
                    event.date.isoformat(),
                    event.start_time.strftime("%H:%M"),
                    event.location[:20],
                    str(len(registrations)),
                    group_info[:30]
                ])
            
            # Create table
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            
            elements.append(table)
            doc.build(elements)
            output.seek(0)
            
            return Response(
                content=output.read(),
                media_type="application/pdf",
                headers={"Content-Disposition": "attachment; filename=event_history.pdf"}
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="reportlab is required for PDF export. Install it with: pip install reportlab")
    
    else:
        raise HTTPException(status_code=400, detail="Format must be 'xlsx' or 'pdf'")

# Export all events with detailed participant information
@app.get("/api/events/export/all")
def export_all_events_with_participants(
    format: str = "xlsx",  # "xlsx" or "pdf"
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Тек администраторлар іс-шараларды экспорттай алады")
    
    from fastapi.responses import Response
    import io
    
    # Get all events with participants
    events = db.query(Event).order_by(Event.date.desc(), Event.start_time.desc()).all()
    
    if format == "xlsx":
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "All Events"
            
            # Headers
            headers = ["Event ID", "Event Title", "Date", "Time", "Location", "Max Participants", "Participant Name", "Participant Email", "Participant Group", "Registration Date"]
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data rows
            for event in events:
                registrations = db.query(EventRegistration).filter(EventRegistration.event_id == event.id).all()
                
                if registrations:
                    for reg in registrations:
                        user = db.query(User).filter(User.id == reg.user_id).first()
                        if user:
                            ws.append([
                                event.id,
                                event.title,
                                event.date.isoformat(),
                                event.start_time.strftime("%H:%M"),
                                event.location,
                                event.max_participants,
                                user.full_name,
                                user.email,
                                user.group or "",
                                reg.created_at.isoformat() if reg.created_at else ""
                            ])
                else:
                    # Event with no participants
                    ws.append([
                        event.id,
                        event.title,
                        event.date.isoformat(),
                        event.start_time.strftime("%H:%M"),
                        event.location,
                        event.max_participants,
                        "No participants",
                        "",
                        "",
                        ""
                    ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return Response(
                content=output.read(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=all_events_participants.xlsx"}
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl is required for XLSX export. Install it with: pip install openpyxl")
    
    elif format == "pdf":
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=letter)
            elements = []
            
            styles = getSampleStyleSheet()
            title = Paragraph("All Events with Participants", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 0.2*inch))
            
            # Group by event
            for event in events:
                # Event header
                event_header = Paragraph(f"<b>Event: {event.title}</b><br/>Date: {event.date} {event.start_time.strftime('%H:%M')}<br/>Location: {event.location}", styles['Normal'])
                elements.append(event_header)
                elements.append(Spacer(1, 0.1*inch))
                
                registrations = db.query(EventRegistration).filter(EventRegistration.event_id == event.id).all()
                
                if registrations:
                    # Prepare participant data
                    data = [["Name", "Email", "Group", "Registered At"]]
                    
                    for reg in registrations:
                        user = db.query(User).filter(User.id == reg.user_id).first()
                        if user:
                            data.append([
                                user.full_name,
                                user.email,
                                user.group or "",
                                reg.created_at.strftime("%Y-%m-%d %H:%M") if reg.created_at else ""
                            ])
                    
                    # Create table
                    table = Table(data)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ]))
                    
                    elements.append(table)
                else:
                    no_participants = Paragraph("No participants registered", styles['Normal'])
                    elements.append(no_participants)
                
                elements.append(Spacer(1, 0.3*inch))
                elements.append(PageBreak())
            
            doc.build(elements)
            output.seek(0)
            
            return Response(
                content=output.read(),
                media_type="application/pdf",
                headers={"Content-Disposition": "attachment; filename=all_events_participants.pdf"}
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="reportlab is required for PDF export. Install it with: pip install reportlab")
    
    else:
        raise HTTPException(status_code=400, detail="Format must be 'xlsx' or 'pdf'")

# Export individual event with participants
@app.get("/api/events/{event_id}/export")
def export_event_with_participants(
    event_id: int,
    format: str = "xlsx",  # "xlsx" or "pdf"
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Тек администраторлар іс-шараларды экспорттай алады")
    
    from fastapi.responses import Response
    import io
    
    # Get event
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Іс-шара табылмады")
    
    # Get participants
    registrations = db.query(EventRegistration).filter(EventRegistration.event_id == event_id).all()
    
    if format == "xlsx":
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Event Participants"
            
            # Event info header
            ws.append(["Event Information"])
            ws.append(["Title:", event.title])
            ws.append(["Date:", event.date.isoformat()])
            ws.append(["Time:", event.start_time.strftime("%H:%M")])
            ws.append(["Location:", event.location])
            ws.append(["Max Participants:", event.max_participants])
            ws.append(["Total Participants:", len(registrations)])
            ws.append([])  # Empty row
            
            # Participant headers
            headers = ["Name", "Email", "Group", "Registered At"]
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[9]:  # Row 9 is the header row
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Participant data
            for reg in registrations:
                user = db.query(User).filter(User.id == reg.user_id).first()
                if user:
                    ws.append([
                        user.full_name,
                        user.email,
                        user.group or "",
                        reg.created_at.isoformat() if reg.created_at else ""
                    ])
            
            if not registrations:
                ws.append(["No participants registered"])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return Response(
                content=output.read(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=event_{event_id}_participants.xlsx"}
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl is required for XLSX export. Install it with: pip install openpyxl")
    
    elif format == "pdf":
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=letter)
            elements = []
            
            styles = getSampleStyleSheet()
            title = Paragraph(f"Event: {event.title}", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 0.2*inch))
            
            # Event info
            event_info = [
                ["Date:", event.date.isoformat()],
                ["Time:", event.start_time.strftime("%H:%M")],
                ["Location:", event.location],
                ["Max Participants:", str(event.max_participants)],
                ["Total Participants:", str(len(registrations))]
            ]
            
            info_table = Table(event_info)
            info_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Participants table
            if registrations:
                data = [["Name", "Email", "Group", "Registered At"]]
                
                for reg in registrations:
                    user = db.query(User).filter(User.id == reg.user_id).first()
                    if user:
                        data.append([
                            user.full_name,
                            user.email,
                            user.group or "",
                            reg.created_at.strftime("%Y-%m-%d %H:%M") if reg.created_at else ""
                        ])
                
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                ]))
                
                elements.append(table)
            else:
                no_participants = Paragraph("No participants registered", styles['Normal'])
                elements.append(no_participants)
            
            doc.build(elements)
            output.seek(0)
            
            return Response(
                content=output.read(),
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=event_{event_id}_participants.pdf"}
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="reportlab is required for PDF export. Install it with: pip install reportlab")
    
    else:
        raise HTTPException(status_code=400, detail="Format must be 'xlsx' or 'pdf'")

# Create default admin user on startup
@app.on_event("startup")
async def create_default_admin():
    db = SessionLocal()
    try:
        admin = db.query(User).filter(User.email == "admin@jihc.kz").first()
        if not admin:
            admin = User(
                email="admin@jihc.kz",
                hashed_password=get_password_hash("admin123"),
                full_name="Администратор",
                role="admin"
            )
            db.add(admin)
            db.commit()
            print("Default admin user created: admin@jihc.kz / admin123")
    except Exception as e:
        print(f"Error creating default admin: {e}")
    finally:
        db.close()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
